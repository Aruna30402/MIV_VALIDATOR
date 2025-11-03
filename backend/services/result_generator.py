"""Module 3: Result Generation Service"""
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Dict, List
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from backend.config import RESULTS_DIR


class ResultGenerator:
    """Generate validation results and downloadable reports"""
    
    def __init__(self):
        """Initialize the result generator"""
        self.results_dir = Path(RESULTS_DIR)
        self.results_dir.mkdir(exist_ok=True)
    
    def create_result_excel(self, original_data: List[Dict], 
                           validation_results: List[Dict],
                           excel_metadata: Dict) -> str:
        """
        Create downloadable Excel file with validation results
        
        Args:
            original_data: Original data from Excel
            validation_results: AI validation results
            excel_metadata: Metadata from Excel processing
            
        Returns:
            Path to generated Excel file
        """
        # Combine original data with validation results
        result_data = []
        
        for orig_item, validation in zip(original_data, validation_results):
            row = orig_item['original_data'].copy()
            
            # Add validation columns
            row['STATUS'] = validation.get('status', 'REVIEW_REQUIRED')
            row['REASON'] = validation.get('reason', 'No reason provided')
            row['CONFIDENCE_SCORE'] = validation.get('confidence_score', 0.70)
            
            result_data.append(row)
        
        # Create DataFrame
        df_results = pd.DataFrame(result_data)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"validation_results_{timestamp}.xlsx"
        filepath = self.results_dir / filename
        
        # Save to Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_results.to_excel(writer, index=False, sheet_name='Validation Results')
            
            # Add summary sheet
            self._create_summary_sheet(writer, validation_results)
        
        # Add hyperlinks to HTTP URLs
        self._add_hyperlinks_to_excel(filepath)
        
        return str(filepath)
    
    def _add_hyperlinks_to_excel(self, filepath: Path):
        """
        Add clickable hyperlinks to HTTP URLs in Excel file
        
        Args:
            filepath: Path to Excel file
        """
        try:
            workbook = load_workbook(filepath)
            worksheet = workbook['Validation Results']
            
            # Pattern to match HTTP/HTTPS URLs
            url_pattern = re.compile(r'https?://[^\s<>"{}|\\^`\[\]]+', re.IGNORECASE)
            
            # Iterate through all cells in the worksheet
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Check if cell contains HTTP URL
                        if url_pattern.search(cell.value):
                            # If cell contains only the URL, make it a hyperlink
                            url_match = url_pattern.match(cell.value.strip())
                            if url_match:
                                url = url_match.group()
                                # Set hyperlink
                                cell.hyperlink = url
                                # Style as hyperlink (blue, underlined)
                                cell.font = Font(color="0563C1", underline="single")
                                # Keep the URL as display text
                                cell.value = url
            
            workbook.save(filepath)
        except Exception as e:
            print(f"Warning: Could not add hyperlinks to Excel: {str(e)}")
            # Continue even if hyperlink addition fails
    
    def _create_summary_sheet(self, writer, validation_results: List[Dict]):
        """
        Create summary statistics sheet
        
        Args:
            writer: Excel writer object
            validation_results: Validation results list
        """
        # Calculate statistics
        total = len(validation_results)
        accepted = sum(1 for r in validation_results if r.get('status') == 'ACCEPTED')
        rejected = sum(1 for r in validation_results if r.get('status') == 'REJECTED')
        review_required = sum(1 for r in validation_results if r.get('status') == 'REVIEW_REQUIRED')
        
        high_confidence = sum(1 for r in validation_results if r.get('confidence') == 'HIGH')
        medium_confidence = sum(1 for r in validation_results if r.get('confidence') == 'MEDIUM')
        low_confidence = sum(1 for r in validation_results if r.get('confidence') == 'LOW')
        
        # Create summary data
        summary_data = {
            'Metric': [
                'Total Images',
                'Accepted',
                'Rejected',
                'Review Required',
                'High Confidence',
                'Medium Confidence',
                'Low Confidence',
                'Acceptance Rate %',
                'Rejection Rate %'
            ],
            'Value': [
                total,
                accepted,
                rejected,
                review_required,
                high_confidence,
                medium_confidence,
                low_confidence,
                f"{(accepted/total*100):.2f}" if total > 0 else "0",
                f"{(rejected/total*100):.2f}" if total > 0 else "0"
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
    
    def create_dashboard_data(self, validation_results: List[Dict]) -> Dict:
        """
        Create data for dashboard display
        
        Args:
            validation_results: Validation results list
            
        Returns:
            Dashboard data dictionary
        """
        total = len(validation_results)
        
        # Status breakdown
        status_counts = {
            'ACCEPTED': 0,
            'REJECTED': 0,
            'REVIEW_REQUIRED': 0
        }
        
        # Confidence breakdown
        confidence_counts = {
            'HIGH': 0,
            'MEDIUM': 0,
            'LOW': 0
        }
        
        for result in validation_results:
            status = result.get('status', 'REVIEW_REQUIRED')
            confidence = result.get('confidence', 'MEDIUM')
            
            if status in status_counts:
                status_counts[status] += 1
            if confidence in confidence_counts:
                confidence_counts[confidence] += 1
        
        return {
            'total': total,
            'status_breakdown': status_counts,
            'confidence_breakdown': confidence_counts
        }
    
    def get_result_file(self, filename: str) -> Path:
        """
        Get path to a result file
        
        Args:
            filename: Name of the result file
            
        Returns:
            Path to the file
        """
        return self.results_dir / filename

